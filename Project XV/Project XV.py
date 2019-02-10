#-------------------------------------------------------------------------------
# Name:        module1
# Purpose:
#
# Author:      user
#
# Created:     19/01/2019
# Copyright:   (c) user 2019
# Licence:     <your licence>
#-------------------------------------------------------------------------------

#import needed object
import os
import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart

def accessXlsxFile():
    """
    Function to access to the xlsx file, return the open file
    """
    #change current directory file
    os.chdir("D:/IIHT/Python/Project/Project XV")
      # open a excel file with .xlsx format, return an object of type openpyxl.workbook.workbook.Workbook
    excelfile = openpyxl.load_workbook('test.xlsx')
    # get the first spreadsheet by name, return an object of type openpyxl.worksheet.worksheet.Worksheet
    sheet1 = excelfile.get_sheet_by_name("test")
    #return the object sheet1
    return sheet1

def createDictionnary(openFile):
    """
    Function to read the open file and create the object latecomer with the name and email for all members behind on their dues
    """
    # get the number of rows in the sheet, return an object of type int
    max_row=openFile.max_row
    # get the number of columns in the sheet, return an object of type int
    max_column=openFile.max_column
    #create the object latecomer of type dict
    latecomer = {}
    # read the whole spreadsheet
    for row_index in range(2, (max_row+1)):
        #condition to compare the value of the last column to None
        if openFile.cell(row=row_index, column=max_column).value == None:
            #creates a key and its value in the object latecomer
            latecomer[openFile.cell(row=row_index, column = 1).value] = openFile.cell(row=row_index, column = 2).value
    #return the object latecomer
    return latecomer

def sendEmail(dictionary):
    """
    functions to send an email to all members behind on their dues
    """
    #connect to the server ,return the object server of type smtplib.SMTP
    server = smtplib.SMTP('smtp.gmail.com', 587)
    #Identify yourself to an SMTP server, return an object of type tuple
    server.ehlo()
    #Put the SMTP connection in TLS (Transport Layer Security) mode, return an object of type tuple
    server.starttls()
    #connect to the email account, return an object of type tuple
    server.login("patdenzoungany@gmail.com", "bermanpat")

    for key,value in dictionary.items():
        #initialise the object to with value, return the object to of type str
        to = value
        #initialise the message to send, return the object mess of type str
        mess = "Mr.{}, we inform you that you have one or more late payments. Please remember to pay them as soon as possible.Thanks for your understanding!".format(key)
        #send the mail, return an object of type dict
        server.sendmail("patdenzoungany@gmail.com", to, mess)
    #close the server, return an object of type tuple
    server.quit()

sheet = accessXlsxFile()
latecomer = createDictionnary(sheet)
sendEmail(latecomer)